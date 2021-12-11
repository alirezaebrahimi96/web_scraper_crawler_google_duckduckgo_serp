from selenium import webdriver
from bs4 import BeautifulSoup
import requests
from fake_useragent import UserAgent
from stem import Signal
from stem.control import Controller
import time
from stem import Signal
from stem.control import Controller
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from functools import wraps
from requests.exceptions import RequestException
from socket import timeout

class Retry(object):
    """Decorator that retries a function call a number of times, optionally
    with particular exceptions triggering a retry, whereas unlisted exceptions
    are raised.
    :param pause: Number of seconds to pause before retrying
    :param retreat: Factor by which to extend pause time each retry
    :param max_pause: Maximum time to pause before retry. Overrides pause times
                      calculated by retreat.
    :param cleanup: Function to run if all retries fail. Takes the same
                    arguments as the decorated function.
    """
    def __init__(self, times, exceptions=(IndexError), pause=1, retreat=1,
                 max_pause=None, cleanup=None):
        """Initiliase all input params"""
        self.times = times
        self.exceptions = exceptions
        self.pause = pause
        self.retreat = retreat
        self.max_pause = max_pause or (pause * retreat ** times)
        self.cleanup = cleanup

    def __call__(self, f):
        """
        A decorator function to retry a function (ie API call, web query) a
        number of times, with optional exceptions under which to retry.

        Returns results of a cleanup function if all retries fail.
        :return: decorator function.
        """
        @wraps(f)
        def wrapped_f(*args, **kwargs):
            for i in range(self.times):
                # Exponential backoff if required and limit to a max pause time
                pause = min(self.pause * self.retreat ** i, self.max_pause)
                try:
                    return f(*args, **kwargs)
                except self.exceptions:
                    if self.pause is not None:
                        time.sleep(pause)
                    else:
                        pass
            if self.cleanup is not None:
                return self.cleanup(*args, **kwargs)
        return wrapped_f

def failed_call(*args, **kwargs):
    """Deal with a failed call within various web service calls.
    Will print to a log file with details of failed call.
    """
    print("Failed call: " + str(args) + str(kwargs))
    # Don't have to raise this here if you don't want to.
    # Would be used if you want to do some other try/except error catching.
    raise RequestException


#Class instance to use as a retry decorator
retry = Retry(times=5, pause=1, retreat=2, cleanup=failed_call,
              exceptions=(RequestException, timeout))

# Function to convert  

def listToString(s): 
    
    # initialize an empty string
    str1 = "" 
    
    # traverse in the string  
    for ele in s: 
        if ele == '\n':
            str1 += '\n'
        str1 += ele  
    
    # return string  
    return str1 

workbook = xlsxwriter.Workbook('result.xlsx')
worksheet = workbook.add_worksheet()

headers = { 'User-Agent': UserAgent().random }


with Controller.from_port(port = 9051) as c:
    c.authenticate("welcome")
    c.signal(Signal.NEWNYM)


proxies = {
    'http': 'socks5h://127.0.0.1:9050',
    'https': 'socks5h://127.0.0.1:9050'
}


PROXY = "127.0.0.1:9050"
firefox_options = webdriver.FirefoxOptions()
firefox_options.add_argument('--proxy-server=%s' % PROXY)
#firefox_options.add_argument('--proxy-server=%s')


driver = webdriver.Firefox(options=firefox_options)



wb = load_workbook('keywords.xlsx')
ws = wb['Sheet1']
 
data = ws.values
# Get the first line in file as a header line
columns = next(data)[0:]
# Create a DataFrame based on the second and subsequent lines of data
df = pd.DataFrame(data, columns=columns)
datafarame = df.values.tolist()
content=[]
for i in range(len(datafarame)):
    content.append(datafarame[i][0])

@retry
def res(url, proxies, headers):
    response=requests.get(url, proxies=proxies, headers=headers, timeout=4)
    return response
l=2
for word in content:
    test = []
    print(word)
    query = word
    links=[]
    for k in [10, 20, 30, 40, 50]:
        url = "http://www.google.com/search?q=" + query + '&start=' + str(k)
        driver.get(url)
        time.sleep(5)
        for i in range(10):
            search = driver.find_elements_by_xpath('/html/body/div[7]/div/div[10]/div[1]/div/div[2]/div[2]/div/div/div['+str(i)+']/div/div/div[1]/a')
            for h in search:
                if 'youtube' not in h.get_attribute('href') and 'google' not in h.get_attribute('href'):
                    links.append(h.get_attribute('href'))
    n = 0
    for i in links:
        text = []
        newlist = []
        url = i
        print(url)
        try:
            
            response = res(url, proxies, headers)
            print(response.status_code)
        
            if response.status_code == 200:
                n += 1
                soup=BeautifulSoup(response.text, 'html.parser')
                for char in ["p"]:
                    parag = soup.findAll(char)
                    nn = 0
                    for j in parag:
                        if '.' in j.text and len(j.text)>70:
                            text.append(j.text)
                            for i in text:
                                if i not in newlist:
                                    newlist.append(i)
                                    newlist.append('\n')
                            newlist.append('\n')
                    passage = listToString(newlist)


                    '''
                    ss = ''
                        
                    n = 0
                    for charac in passage:
                        if charac == '.':

                            if n == 4:
                                passage = passage.replace(str(charac),'\n')
                                n = 0
                                ss+='.'
                                ss+='\n'
                            n+=1
                        ss += charac
                    for charac in ss:
                        ss = ss.replace('\n.','\n')
                    '''



                    file = open(str(word),"a")
                    file.write(passage)
                    file.close()
                    test.append(passage)

        except Exception as e:
            print(e)
    test2 = listToString(test)
    worksheet.write('A1', 'keywords')
    worksheet.write('B'+str(l), test2)
    worksheet.write('A'+str(l), word)
    l += 1
driver.close()
workbook.close()